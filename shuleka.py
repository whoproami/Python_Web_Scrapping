from bs4 import BeautifulSoup
import requests
from openpyxl import Workbook

html= requests.get("https://www.sulekha.com/loans/Kolkata");

wb =  Workbook()

delhi_sheet1=wb.active
# print(html.text)
soup = BeautifulSoup(html.text,'lxml')
liste = soup.find_all('div');
i=1;
for link in liste:
    i_d = link.get('id');
    tags=link.find(class_="tags");
    if i_d !=None:
       if i_d.startswith("businessdiv"):
            anc = link.find('h3');
            print(anc.text.strip());
            spn = link.find('span');
            print(spn.text.strip());
            acr = link.find_all("a");
            for lin in acr:
              i_d2 =lin.get('id');
              if i_d2:
                  if  i_d2.startswith("callIcon"):
                     print(lin.get('href').strip())
                     delhi_sheet1.cell(i,1).value=anc.text.strip()
                     delhi_sheet1.cell(i,2).value=spn.text.replace(" ","").strip()
                     delhi_sheet1.cell(i,3).value=lin.get('href').strip()
                  if i_d2.startswith("businessname"):
                     print(lin.get('href').strip())
                     delhi_sheet1.cell(i+1,4).value=lin.get('href').strip()
                  if tags:
                     tgs = tags.find('span').text.strip()
                     delhi_sheet1.cell(i+1,5).value=tgs


                      
                  
                  i+=1

wb.save('kolkata.xls')

             
                
            
            
            
            
            
        
    

# print(soup.prettify())
# f = open("file.txt" , 'w');
# f.write(soup)
# f.close();