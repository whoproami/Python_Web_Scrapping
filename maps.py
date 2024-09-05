from selenium import webdriver
from bs4 import BeautifulSoup
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
import re
from openpyxl import Workbook

wb =  Workbook()

sheet1=wb.active

def extr_no():
    options = webdriver.ChromeOptions()
    options.add_argument("disable-infobars")
    options.add_argument("start-maximized")
    options.add_argument("disable-dev-shm-usage")  # For Linux issues
    options.add_argument("no-sandbox")
    options.add_experimental_option("excludeSwitches", ['enable-automation'])
    options.add_argument("disable-blink-features=AutomationControlled")

    service = Service("/home/priya/Learning/selenium/Lending Budha/old _files/chromedriver copy")

    driver = webdriver.Chrome(service=service, options=options)

    print("Enter the City Name(Please press Enter After it):")
    city = input()    
    
    driver.get(f"https://www.google.com/maps/search/loan+dsa+in+{city}")

    html = driver.page_source

    driver.quit()

    soup = BeautifulSoup(html, "lxml")

    list_super_div = soup.find_all("div",attrs={'class':"Z8fK3b"});
    list_div = soup.find_all("div", attrs={'aria-label': f'Results for loan dsa in {city}'})
    list_name = soup.find_all("div",attrs={'class':"qBF1Pd"})
    


    phone_numbers = []

    i = 1;
    for j in list_super_div:
        try :
         listno = j.find("span",attrs={'class':"UsdlK"})
         listname = j.find("div",attrs={'class':"qBF1Pd"})
         listdata = j.find("div",attrs={'class':"W4Efsd"})
         print(listname.text,"+>",listno.text)

         sheet1.cell(i,1).value=listname.text
         sheet1.cell(i,2).value=listno.text
         i+=1
        except :
           print("mumber not given")
        
    
    wb.save(f'{city}.xls')
 


phone_numbers = extr_no()
# print(phone_numbers)
